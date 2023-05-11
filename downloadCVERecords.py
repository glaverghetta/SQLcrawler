import re
import requests
from bs4 import BeautifulSoup
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font

def get_url(begin:date):
    url = "https://nvd.nist.gov/vuln/search/results?form_type=Advanced&results_type=overview&search_type=all&cwe_id=CWE-89&pub_start_date={start}&pub_end_date={end}&orderBy=publishDate&orderDir=asc"
    end = begin + timedelta(days=119)  # Bugs out with 120 from August 29-December 31, which is the reported max

    if (end.year > begin.year):
        end = date(year=begin.year, month=12, day=31)

    start_string = f"{begin.month:02d}/{begin.day:02d}/{begin.year}"
    end_string = f"{end.month:02d}/{end.day:02d}/{end.year}"
    return (url.format(start=start_string, end=end_string), end)

def panic(s, cve):
    print("Error: ", )
    print(s)
    print(cve)
    exit(-1)

def getNumberAndLink(cve):
    th = cve.select("th")
    if len(th) != 1:
        panic("Found too many headers in getNumberAndLink", cve)
    a = th[0].select("a")
    if len(a) != 1:
        panic("Found too many links in getNumberAndLink", cve)
    
    return a[0].text, "https://nvd.nist.gov" + a[0].get("href") 

def getDescriptionAndDate(cve):
    td = cve.select("td")
    if len(td) != 2:
        panic("Found too many rows in getDescriptionAndDate", cve)
    p = td[0].select("p")
    if len(p) != 1:
        panic("Found too many paragraphs in getDescriptionAndDate", cve)
    span = td[0].select("span")
    if len(span) != 1:
        panic("Found too many spans in getDescriptionAndDate", cve)
    
    return (p[0].text, span[0].text.split(";")[0])

def getScore(cve):
    td = cve.select("td")
    if len(td) != 2:
        panic("Found too many rows in getScore", cve)
    a = td[1].select("a")
    
    return (a[0].text, "https://nvd.nist.gov" + a[0].get("href"))


def saveCVE(cve, sheet, row):
    number, link = getNumberAndLink(cve)
    description, date = getDescriptionAndDate(cve)
    score, scoreLink = getScore(cve)

    sheet[f"A{row}"].hyperlink = link
    sheet[f"A{row}"].value = number
    sheet[f"A{row}"].style = "Hyperlink"

    sheet[f"B{row}"] = date
    sheet[f"C{row}"] = description

    sheet[f"D{row}"].hyperlink = scoreLink
    sheet[f"D{row}"].value = score
    sheet[f"D{row}"].style = "Hyperlink"
    background, font = getScoreColor(score)
    sheet[f"D{row}"].font = Font(color=font)
    sheet[f"D{row}"].fill = PatternFill("solid", fgColor=background)

def getScoreColor(score):
    if "CRITICAL" in score:
        return "ff0000", "000000"
    if "HIGH" in score:
        return "d9534f", "000000"
    if "MEDIUM" in score:
        return "ec971f", "000000"
    if "LOW" in score:
        return "f2cc0c", "000000"
    return 'ff00ff00', "000000"

current = date(year=2022, month=8, day=31)
end = date(year=2023, month=1, day=1)

filename = "output.xlsx"

workbook = Workbook()
sheet = workbook.active

row = 1

while current < end:
    url, endDate = get_url(current)
    pageNumber = 1

    while url is not None:
        print(f"Retrieving page {pageNumber} for {current} to {endDate}")
        page = requests.get(url)

        soup = BeautifulSoup(page.content, "html.parser")

        cves = soup.find_all("tr", attrs={"data-testid": re.compile(r'vuln-row-[0-9]+')})

        if len(cves) == 0:
            panic(f"No CVEs found for {url}!", "")

        for cve in cves:
            saveCVE(cve, sheet, row)
            row += 1
        
        nextPage = soup.select("a[data-testid='pagination-link-page->']")
        if len(nextPage) == 2:  # Has one at top and bottom of page
            url = "https://nvd.nist.gov" + nextPage[0].get("href")
            pageNumber += 1
        else:
            url = None
    current = endDate + timedelta(days=1)

workbook.save(filename=filename)